# openpyxl 설치: pip install openpyxl
# Python에서 엑셀을 쉽게 다룰 수 있도록 도와주는 라이브러리
# 반드시 확장자가 .xlsx 여야한다.
from openpyxl import Workbook

#시간 가져오기
from datetime import datetime


today = datetime.today().strftime("%Y%m%d")
print(today)

#엑셀 세팅
wb = Workbook()
ws1 = wb.active

#sheet이름
ws1.title = today

#행 추가 
ws1.append(["user Name", "Real Name", "Company", "Host Name", "Last Accessed Date"])

#행 추가 
ws1.append(['1','2','3','4','5'])

#파일이름
wb.save(filename = today + '.xlsx')

# #엑셀 저장하기
# wb.save('파일이름.xlsx')
# wb.save(r'저장경로\파일이름.xlsx')
# #문자열 앞에 r은 다른걸로 인식하지말고 단순히 문자열로 인식해라 라는 뜻.


#엑셀 파일을 새로 만드는 방법
# wb = openpyxl.Workbook() #새로운 Workbook 객체 생성

# #Workbook 객체를 실제 파일에 저장하기 위해 .save(파일명) 작성
# wb.save("파일명") 
# #저장 할 위치를 정해주고 싶다면 경로 설정
# wb.save(r'C:\경로')

#기존에 있던 엑셀 파일을 불러오는 방법
# #path 변수에 경로 저장
# path= r'C:\경로'
# #미리 만들어놓은 엑셀파일을 불러옴
# wb= openpyxl.load_workbook(path)


